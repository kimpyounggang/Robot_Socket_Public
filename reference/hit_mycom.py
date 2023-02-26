from cProfile import label
from ctypes import resize
from multiprocessing import Process
from msilib.schema import ComboBox
from sqlite3 import register_converter
import sys,os,time,clipboard,pywinauto,shutil,subprocess,pyautogui
from datetime import date
from tkinter import OFF, Toplevel
from turtle import color
from datetime import date
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt as qt
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import pandas as pd
import matplotlib.style as mplstyle
from pywinauto.application import Application


matplotlib.use('Qt5Agg')
mplstyle.use('fast')
matplotlib.rcParams['agg.path.chunksize']=10000


class sub(QDialog):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('target_n 매크로')
        self.setGeometry(300, 300, 1500, 300) #창의크기 3번째 가로  4번째 세로
        
        layout = QVBoxLayout()
        layout1 = QHBoxLayout()
        layout2 = QHBoxLayout()
        layout_btn = QVBoxLayout()
        
        #라벨
        self.label_1 = QLabel('변경할 폴더의 주소 = \n사진숫자 검색 주쇼 = ')
        layout1.addWidget(self.label_1 )
        #입력
        self.insert_1 = QLineEdit()
        layout1.addWidget(self.insert_1) #파일주소 입력

        #라벨
        self.label_2 = QLabel('이동할 폴더의 주소 = ')
        layout2.addWidget(self.label_2)
        #입력
        self.insert_2 = QLineEdit()
        layout2.addWidget(self.insert_2)

        # 버튼 정의
        self.btnRun_1 = QPushButton("target 변환", self)
        layout_btn.addWidget(self.btnRun_1)
        self.btnRun_1.clicked.connect(self.btnRun_1_clicked_all_str)	# 클릭 시 실행할 function
        # 버튼 정의
        self.btnRun_4 = QPushButton("폴더 사진숫자 검사", self)
        layout_btn.addWidget(self.btnRun_4)
        self.btnRun_4.clicked.connect(self.btnRun_4_clicked_spe_str)	# 클릭 시 실행할 function
        
        layout1.addLayout(layout_btn )
        #라벨
        self.label_4 = QLabel('변환 시작번호 부여 : ')
        layout1.addWidget(self.label_4 )
        #입력
        self.insert_4 = QLineEdit()
        layout1.addWidget(self.insert_4)
        # 버튼 정의
        self.btnRun_5 = QPushButton("target 변환 시작번호", self)
        layout1.addWidget(self.btnRun_5,stretch=1)
        self.btnRun_5.clicked.connect(self.btnRun_5_clicked_spe_str)	# 클릭 시 실행할 function


        # 버튼 정의
        self.btnRun_2 = QPushButton("이동", self)
        layout2.addWidget(self.btnRun_2)
        self.btnRun_2.clicked.connect(self.btnRun_2_clicked_spe_str)	# 클릭 시 실행할 function
        
        
        layout.addLayout(layout1)
        layout.addLayout(layout2)
        
        self.setLayout(layout)
        self.show()
    #     self.init_

    # def init_(self):
    #     dir = Main.log_read_line(2)


    def btnRun_1_clicked_all_str(self):#변환 시작
        folders = []
        filedir = self.insert_1.toPlainText() # 생산되는 폴더의 주소
        folders = os.listdir(filedir) #현재 폴더 목록 불러와서 리스트에 넣는
        
        for i,number in zip(folders,range(1,len(folders)+1)):
            os.rename(f"{filedir}/{i}", f"{filedir}/target_{number}")
            
        return None
    
    def btnRun_2_clicked_spe_str(self):#이동만
        folders = []
        filedir = self.insert_1.toPlainText() # 생산되는 폴더의 주소
        folders = os.listdir(filedir) #현재 폴더 목록 불러와서 리스트에 넣는다
        movedir = self.insert_2.toPlainText() # 붙여넣을 폴더의 주소
        
        #본 폴더 이동
        for i in folders:
            shutil.move(f"{filedir}/{i}", f"{movedir}/{i}")
        return None
    
    def btnRun_4_clicked_spe_str(self):#폴더 사진숫자 검사
        folders = []
        filedir = self.insert_1.toPlainText() # 생산되는 폴더의 주소
        folders = os.listdir(filedir) #현재 폴더 목록 불러와서 리스트에 넣는다
        
        for i in folders:
            if len(os.listdir(f"{filedir}/{i}"))!=46:
                QMessageBox.information("사진 숫자 부족",  f"{i}폴더의 사진이 부족합니다.")
        
        QMessageBox.information("사진숫자 검색결과", "이상없음")
        
        return None
            
            
    def btnRun_5_clicked_spe_str(self):#변환 시 시작번호 부여
        start_num = int(self.insert_4.toPlainText())
        folders = []
        filedir = self.insert_1.toPlainText() # 생산되는 폴더의 주소
        folders = os.listdir(filedir) #현재 폴더 목록 불러와서 리스트에 넣는다
        
        try:
            for i,number in zip(folders,range(start_num,len(folders)+start_num)):
                os.rename(f"{filedir}\\{i}", f"{filedir}\\target_{number}")
        except:
            for i,number in zip(folders,range(1,len(folders)+1)):
                os.rename(f"{filedir}\\{i}", f"{filedir}\\{number}")
            start_num = int(self.insert_4.toPlainText())
            folders = []
            filedir = self.insert_1.toPlainText() # 생산되는 폴더의 주소
            folders = os.listdir(filedir) #현재 폴더 목록 불러와서 리스트에 넣는다
            for i,number in zip(folders,range(start_num,len(folders)+start_num)):
                os.rename(f"{filedir}\\{i}", f"{filedir}\\target_{number}")
                
        return None

class MplCanvas(FigureCanvasQTAgg):
    def __init__(self):
        fig = Figure()
        self.axes = fig.add_subplot(111)
        self.axes.axis([-15, 15, -15, 15])
        self.axes.grid()   
        super(MplCanvas, self).__init__(fig)


class Main(QDialog):

    def __init__(self):
        super().__init__()
#################전역변수#################
        self.take_point = []
        self.mode = self.set_reference_x = self.finder = self.finds =self.file_name = ''
        self.file_dir_r1 =self.file_dir_r2 =self.file_dir_r3 =self.file_dir_r4 =self.file_dir_r5= ''
        self.r_u_search = self.set_data = self.number= 0
        self.r = [0,0,0,0,0]
        self.modech = 0
        self.aver_logic=[]
        # GUI
        self.Round = [0,0,0,0,0,0]
        self.R_Create = [0,0,0,0,0,0]
        self.r_btn = [0,0,0,0,0,0]
        self.R_Creat_ = [0,0,0,0,0,0]
        self.insert_routine = [0,0,0,0,0,0]
        ## df
        today = date.today() #오늘 날짜를 불러오는 함수
        self.today_ = today.strftime('%y%m%d')

############################################
        self.init_ui()

    def init_ui(self):
        self.sc = MplCanvas()
        self.setGeometry(300, 100, 800, 800)
        #self.setFixedSize(1200,900)
        self.setFixedSize(2600,1400)
        toolbar = NavigationToolbar(self.sc, self)
        layout = QVBoxLayout()
        layout_equation_button = QHBoxLayout()
        layout_search = QHBoxLayout()
        layout_rud_btn = QHBoxLayout()
        layout_create = QHBoxLayout()
        layout_create_sqr = QHBoxLayout()
        layout_insert_routine = QHBoxLayout()
        layouts = QHBoxLayout()

######################콤보리스트 생성####################
        self.combo_list = QComboBox(self)
        points = []
        for i in range(0,47):
            points.append(f"Point{i+1}")
        self.combo_list.addItems(points)
        self.combo_list.setGeometry(500,100,170,30)#x,y, x크기,y크기
        self.combo_list.currentIndexChanged.connect(self.combo_fun)
        
#################위젯정의############################
        for i in range(1,6):
            # csv등록 버튼
            self.Round[i] = QPushButton(f"File {i}")#버튼생성
            if i == 1: self.Round[i].clicked.connect(self.Round_1)#함수연결
            if i == 2: self.Round[i].clicked.connect(self.Round_2)
            if i == 3: self.Round[i].clicked.connect(self.Round_3)
            if i == 4: self.Round[i].clicked.connect(self.Round_4)
            if i == 5: self.Round[i].clicked.connect(self.Round_5)
            layout_equation_button.addWidget(self.Round[i]) # 위젯등록
            # 편차 그래프 버튼
            self.R_Create[i] = QPushButton(f"R{i} Create")#버튼생성
            if i == 1: self.R_Create[i].clicked.connect(self.r1_create)#함수연결
            if i == 2: self.R_Create[i].clicked.connect(self.r2_create)
            if i == 3: self.R_Create[i].clicked.connect(self.r3_create)
            if i == 4: self.R_Create[i].clicked.connect(self.r4_create)
            if i == 5: self.R_Create[i].clicked.connect(self.r5_create)
            layout_create.addWidget(self.R_Create[i]) # 위젯등록
            # 포인트 그래프 체크박스
            self.r_btn[i] = QCheckBox(f"{i}rud")#체크박스생성
            if i == 1: self.r_btn[i].stateChanged.connect(self.draw_round1)#함수연결
            if i == 2: self.r_btn[i].stateChanged.connect(self.draw_round2)
            if i == 3: self.r_btn[i].stateChanged.connect(self.draw_round3)
            if i == 4: self.r_btn[i].stateChanged.connect(self.draw_round4)
            if i == 5: self.r_btn[i].stateChanged.connect(self.draw_round5)
            layout_rud_btn.addWidget(self.r_btn[i]) # 위젯등록
            # 개개 편차 텍스트
            self.R_Creat_[i] = QTextBrowser() #텍스트생성
            self.R_Creat_[i].append('Devitaion')
            layout_create_sqr.addWidget(self.R_Creat_[i]) # 위젯등록
            # 개개 데이터 루틴 시작
            self.insert_routine[i] = QPushButton(f"{i}_Open")#루틴시작 버튼
            if i == 1: self.insert_routine[i].clicked.connect(self.open_r1)#함수연결
            if i == 2: self.insert_routine[i].clicked.connect(self.open_r2)
            if i == 3: self.insert_routine[i].clicked.connect(self.open_r3)
            if i == 4: self.insert_routine[i].clicked.connect(self.open_r4)
            if i == 5: self.insert_routine[i].clicked.connect(self.open_r5)
            layout_insert_routine.addWidget(self.insert_routine[i]) # 위젯등록
        
        self.equation = QLineEdit() #주소 입력
        self.datalabel = QLabel('Search Data') # 검색한 데이터
        search_btn = QPushButton("Search csv") # 데이터 검색
        ###
        reset_btn = QPushButton("Rset") ## 리셋버튼
        reset_btn.setStyleSheet('QPushButton { color: red;}')
        self.mod = QPushButton("Mode Change") #
        self.mod.setStyleSheet('QPushButton { color: blue;}')
        ###
        self.status = QPushButton('INIT')
        self.status.setStyleSheet('QPushButton { color: red;}')
        ###
        reset_check = QPushButton("Reset_check") #체크해제
        
##################위젯 함수연결############################
        search_btn.clicked.connect(self.search_xl) #검색버튼
        ###
        reset_btn.clicked.connect(self.reset) #리셋버튼
        ###
        reset_check.clicked.connect(self.reset_check) #체크해제 버튼
        self.mod.clicked.connect(self.modechange)
        self.status.clicked.connect(self.set_window)
##################위젯등록################
        layout_search.addWidget(self.equation) #주소 입력
        layout_search.addWidget(self.datalabel) #검색한 파일
        layout_search.addWidget(search_btn) #주소 검색
        ##
        layout_equation_button.addWidget(reset_btn)   #리셋버튼
        ##
        layout_insert_routine.addWidget(self.status)
        ##
        layout_rud_btn.addWidget(toolbar)  #툴바
        layout_rud_btn.addWidget(reset_check) #체크해제
        layout_rud_btn.addWidget(self.combo_list) #포인트
        ##
        layout_create.addWidget(self.mod) #모드선택
        ##
###################레이어 등록#################
        layout.addLayout(layout_search,stretch=1)
        layout.addLayout(layout_equation_button,stretch=1)
        layout.addLayout(layout_create,stretch=1)
        layout.addLayout(layout_insert_routine,stretch=1)
        layout.addLayout(layout_rud_btn,stretch=1)
        layout.addWidget(self.sc,stretch=12)  ##그래프
        layouts.addLayout(layout,stretch=6)
        layouts.addLayout(layout_create_sqr,stretch=4)
        
###############원 + 중심선 생성#############
        # draw_circle = plt.Circle((0, 0), 10, color='black',linewidth=1,fill=False,label='Safety Zone')
        # self.sc.axes.add_artist(draw_circle)
        self.sc.axes.hlines(0, -30, 30,  linestyle='-', linewidth=2) #hlines 수평선 (y)
        self.sc.axes.vlines(0, -30, 30,  linestyle='-', linewidth=2)
        self.sc.draw()
        
##############최종 레이어 등록#############
        self.setLayout(layouts)
        self.show()
        self.init_read()
        #self.init_loop()
    def set_window(self):
        sub().exec_()

    def init_read(self):
        if getattr(sys, 'frozen', False):
            self.nowdir = os.path.dirname(os.path.realpath(sys.executable))
        elif __file__:
            self.nowdir = os.path.dirname(__file__)
        # print(self.nowdir)
        # self.nowdir_files = (os.listdir(self.nowdir))
        # print(self.nowdir_files)
        # #self.nowdir = os.path.dirname(os.path.abspath(__file__))
        # self.nowdir = os.path.dirname(os.path.realpath(sys.executable))
        self.nowdir_files = (os.listdir(self.nowdir))

        befor_dir,befor_csv_name,befor_csv_dir = self.log_read()
        if befor_dir != '0\n':self.equation.setText(befor_dir)
        
         
        for i in range(0,5):
            ##위젯 이름 설정
            if befor_csv_name[i] != '0': self.Round[i+1].setText(befor_csv_name[i])
            ## 데이터 설정
            if befor_csv_dir[i] != '0': setattr(self,f'df{i+1}',pd.read_csv(befor_csv_dir[i], header=0,nrows=46,encoding='CP949'))

    # def vision_data_read(self):
    #     if txtlist.index(i) >= 15 and txtlist.index(i)%4==3:
            


    def data_verif(self,df):
        with open(f'{self.nowdir}\\data_log{self.create_number-1}.txt','r',encoding="utf8") as read_txt:
            txtlist1 = read_txt.readlines()
        with open(f'{self.nowdir}\\data_log{self.create_number-2}.txt','r',encoding="utf8") as read_txt:
            txtlist2 = read_txt.readlines()
        if txtlist1[1:] == txtlist2[1:]:
            os.remove(f'{self.nowdir}\\data_log{self.create_number-1}.txt')
            self.create_number = self.create_number-1
            self.log_write(17,f'{self.create_number}')
        else:
            self.data_to_xl(df)
            
    def data_to_xl(self,df):
        dir = f'{self.equation.text()}\\test.xlsx'
        wb=load_workbook(dir)
        name = wb.sheetnames
        sheet = wb[name[2]]
        wb.copy_worksheet(sheet)
        sheet.title = f'{self.today_}_{self.create_number}'
        for i in range(0,46):
            for i1 in range(0,len(df.columns)-1):
                sheet.cell(row=26+i, column=4+i1).value = df.loc[i].iloc[i1]
        wb.save(dir)

    
    def data_write(self,data,df,r_how):
        dev_data = []
        try:
            if 'data_log.txt' and 'log.txt' in self.nowdir_files:
                with open(f'{self.nowdir}\\data_log.txt','r',encoding="utf8") as read_txt:
                    txtlist = read_txt.readlines()
                with open(f'{self.nowdir}\\data_log{self.create_number}.txt', 'w',encoding="utf8") as read_txt:
                        self.create_number = self.create_number+1
                        self.log_write(17,f'{self.create_number}')
                        i2 = 0
                        for i in txtlist:
                            if txtlist.index(i) == 0:
                                    read_txt.write(f'   PROC NONSTOP_1200_ACC6050_{self.today_}_{self.create_number}()')
                            elif txtlist.index(i) >= 15 and txtlist.index(i)%4==3:
                                # if data[i2]<0:
                                #     data[i2]= -data[i2]
                                if data[i2]==0:
                                    data[i2] = 5
                                read_txt.write(f'    TriggIO vision_on,{abs(data[i2])}\\DOp:=do01_VisionTrigg,1;\n')
                                i2 = i2+1
                            else: 
                                read_txt.write(i)

                if self.create_number>1:
                    self.data_verif(df)
                else: 
                    self.data_to_xl(df)
            if r_how == 2:
                self.log_write(19,self.create_number)
            if r_how == 4:
                self.log_write(20,self.create_number)
            if r_how == 6:
                self.log_write(21,self.create_number)
            if r_how == 8:
                self.log_write(22,self.create_number)
            if r_how == 10:
                self.log_write(23,self.create_number)
        except:pass

        return data

    def log_check(self):
        # 자리수 검사
        with open(f'{self.nowdir}\\log.txt','r') as read_txt:
            txtlist = read_txt.readlines()
        if txtlist[0] != 'befor\n' or txtlist[1] != 'befor_dir\n' or txtlist[3] != 'befor_csv_name\n' or txtlist[9] != 'befor_csv_dir\n' or txtlist[15] != 'now\n':
            wrong_data = True
            self.create_log() #초기화
            QMessageBox.warning(self, "Error code 02", "Wrong Data Log")
        return wrong_data

    def log_write(self,write_index,insert_data):
        try:
            if 'log.txt' in self.nowdir_files:
                with open(f'{self.nowdir}\\log.txt','r') as read_txt:
                    txtlist = read_txt.readlines()
                with open(f'{self.nowdir}\\log.txt', 'w') as read_txt:
                    try:
                        for i, txt in enumerate(txtlist,start=0):
                            if i == write_index:
                                read_txt.write(f'{insert_data}\n')
                            else:
                                read_txt.write(txt)
                    except:pass
            # 자리수 검사
                self.log_check()
        except:pass

    def log_read_line(self,read_index):
        try:
            if 'log.txt' in self.nowdir_files:
                with open(f'{self.nowdir}\\log.txt','r') as read_txt:
                    txtlist = read_txt.readlines()
                    for i, txt in enumerate(txtlist,start=0):
                            if i == read_index:
                                read_data = txt
            # 자리수 검사
                self.log_check()
        except:pass
        return read_data


    def log_read(self):
        # 로그있음
        if 'log.txt' in self.nowdir_files:
            with open(f'{self.nowdir}\\log.txt','r') as read_txt:
                txtlist = read_txt.readlines()
                befor_csv_name = []
                befor_csv_dir = []
                befor_dir = '0\n'
                if txtlist[2] != '0\n':  # befor dir
                    res=txtlist[2]
                    befor_dir = res.replace('\n', '')
                for i in range(4,9):  #[4] ~ [8] ->  befor_csv_name 1 2 3 4 5
                        res = txtlist[i]
                        befor_csv_name.append(res.replace('\n', ''))
                for i in range(10,15):  #[4] ~ [8] ->  befor_csv_dir 1 2 3 4 5
                        res = txtlist[i]
                        befor_csv_dir.append(res.replace('\n', ''))
                res = txtlist[17]
                res.replace('\n', '')
                self.create_number = int(res)
        ## 로그파일이 없음
        if 'log.txt' not in self.nowdir_files:
            self.create_log()
        return befor_dir,befor_csv_name,befor_csv_dir

####################### 기본 데이터 덮어씌움 생성 ##################
    def create_log(self):
        with open(f'{self.nowdir}\\log.txt', 'w') as read_txt:
            read_txt.write('befor\n') #0
            read_txt.write('befor_dir\n') #1
            read_txt.write('0\n') #2
            read_txt.write('befor_csv_name\n') #3
            for i in range(0,5):
                read_txt.write('0\n') #4~8
            read_txt.write('befor_csv_dir\n')
            for i in range(0,5):
                read_txt.write('0\n') #9~14
            read_txt.write('now\n')
            read_txt.write('create_number\n') #16(실시회차)
            read_txt.write('0\n') #17
            read_txt.write('last_data_number\n') #18(실시회차)
            read_txt.write('0\n') #19
            read_txt.write('0\n') #20
            read_txt.write('0\n') #21
            read_txt.write('0\n') #22
            read_txt.write('0\n') #23


    def reset_check(self):
        for i in range(1,6):
            self.r_btn[i].setChecked(False)
    
    def Round_1(self):
        try:
            if self.r[0] == 0:
                self.Round[1].setText(self.datalabel.text())
                self.file_name = self.datalabel.text()
                file_dir_r1 = f'{self.finder}\\{self.file_name}'
                self.df1 = pd.read_csv(file_dir_r1, header=0,nrows=47,encoding='CP949')
                self.r[0] = 1
                self.log_write(4,self.file_name)
                self.log_write(10,file_dir_r1)
                # 시트 번호 지정(첫 번째는 0, 두 번째는 1, ...)
                # header : 시작 행 위치 지정 nrows : 끝나는 행 지정 usecols : 가져올 열 수동 지정
        except:pass
    def Round_2(self):
        try:
            if self.r[1] == 0:
                self.file_name = self.datalabel.text()
                self.Round[2].setText(self.datalabel.text())
                file_dir_r2 = f'{self.finder}\\{self.file_name}'
                self.df2 = pd.read_csv(file_dir_r2, header=0,nrows=47,encoding='CP949')
                self.r[1] = 1
                self.log_write(5,self.file_name)
                self.log_write(11,file_dir_r2)
        except:pass
    def Round_3(self):
        try:
            if self.r[2] == 0:
                self.file_name = self.datalabel.text()
                self.Round[3].setText(self.datalabel.text())
                file_dir_r3 = f'{self.finder}\\{self.file_name}'
                self.df3 = pd.read_csv(file_dir_r3, header=0,nrows=47,encoding='CP949')
                self.r[2] = 1
                self.log_write(6,self.file_name)
                self.log_write(12,file_dir_r3)
        except:pass
    def Round_4(self):
        try:
            if self.r[3] == 0:
                self.file_name = self.datalabel.text()
                self.Round[4].setText(self.datalabel.text())
                file_dir_r4 = f'{self.finder}\\{self.file_name}'
                self.df4 = pd.read_csv(file_dir_r4, header=0,nrows=47,encoding='CP949')
                self.r[3] = 1
                self.log_write(7,self.file_name)
                self.log_write(13,file_dir_r4)
        except:pass
    def Round_5(self):
        try:
            if self.r[4] == 0:
                self.file_name = self.datalabel.text()
                self.Round[5].setText(self.datalabel.text())
                file_dir_r5 = f'{self.finder}\\{self.file_name}'
                self.df5 = pd.read_csv(file_dir_r5, header=0,nrows=47,encoding='CP949')
                self.r[4] = 1
                self.log_write(8,self.file_name)
                self.log_write(14,file_dir_r5)
        except:pass
        
    def search_xl(self):
        try:
            self.finder =dir_data= self.equation.text()
            self.log_write(2,dir_data)
            self.finds = os.listdir(self.finder)
            res = []
            for i in self.finds:
                if '~$' in i:
                    self.finds.remove(i) # 현재 실행중인 엑셀 임시파일 제외하고 검색
                if '.csv' in i:
                    res.append(i)
            for i in range(0,6):
                if self.set_data == i:
                    if len(res) > self.r_u_search:
                        self.datalabel.setText(f'{res[self.r_u_search]}')
                        self.datalabel.repaint()
                        self.r_u_search = self.r_u_search+1
                    else:
                        self.r_u_search = 0
                        self.datalabel.setText(f'{res[self.r_u_search]}')
                        self.datalabel.repaint()
                        self.r_u_search = self.r_u_search+1
        except:pass
    
    def reset(self):
        for i in range(0,5):
            self.r[i] = 0
            setattr(self,f'df{i+1}',0)
            self.Round[i+1].setText(f'File {i+1}')
        for i1 , i2 in zip(range(4,9),range(10,15)):
            self.log_write(i1,'0')
            self.log_write(i2,'0')
            
        self.datalabel.setText('Search Data')
        self.datalabel.repaint()
        
    def combo_fun(self):
        ## df 0 -> reference
        try:
            points = []
            for i in range(0,46): ## 포인트 세팅. 46 포인트까지 있어야함
                points.append(f"Point{i+1}")
            ##포인트별 
            for i in points:
                if self.combo_list.currentText() == i:
                    self.number = points.index(i) # i = 0
                    self.setpoint = self.df.loc[self.number]
                    self.set_reference_x = self.setpoint.iloc[0:1] #레퍼런스 포인트 설정
                    self.take_point = self.setpoint[1:]
                    #타겟 1부터 끝 타겟까지############################################################
                    for i in range(0,len(self.take_point)):
                        self.take_point[i] = self.take_point[i]-1500
                    #셋 레퍼런스 한 뒤에, 다른 지점을 레퍼런스로 빼야됨. 점 - 레퍼런스 이렇게.
        except:pass
    
    def maindraw(self,r_w):
        res=0
        self.aver_logic = []
        for i,i_n in zip(self.take_point,range(1,len(self.take_point))):
            self.sc.axes.scatter(i,0,c='black',s=10)  ##여러 산점도
            self.sc.axes.text(i,0+0.2, f'{i_n}',fontdict={'size': 15}) #산점도 레이블
            if i !=-1500:
                res = res+i
                self.aver_logic.append(i)

        res_aver = res/len(self.aver_logic) #평균좌표
        self.sc.axes.scatter((res_aver),0,c='black',s=100) #평균점
        #farpix_x -> 머리막대가 도착하는 x지점 
        far_pix = round(-float(self.set_reference_x-1500-res_aver),2) 
        #평균지점 -> 레퍼런스 지점
        self.sc.axes.annotate('', xytext=(res_aver, -r_w), xy=(self.set_reference_x-1500, -r_w), xycoords='data',
        arrowprops=dict(arrowstyle='->', color='red', lw=3))
        #텍스트 지정
        self.sc.axes.annotate(f'R_{r_w} Fix X pixel = {far_pix}px\nFix X mm = {round(far_pix*0.0725,2)}mm',
                            xy=(self.set_reference_x-1500, -r_w), ha='center', va='baseline', fontsize=15)
        
        
        # val_aver  = far_pix #평균오차
        # print(val_aver)
        val_point = self.take_point.values.tolist()
        val_point.pop()
        val = []
        for i in val_point:
            if i != -1500:
                val.append(i)
        val_max_min = max(val) - min(val) 
        val_max = round(max(val),2) #최대오차    
        val_min = round(min(val),2)   #최소오차
        val_sqr = round(np.std(val),2) #표준편차

        self.sc.axes.annotate(f'max-min : {val_max_min}px , {round(val_max_min*0.0725,2)}mm\n val_max : {val_max}px , {round(val_max*0.0725,2)}mm\n val_min : {val_min}px , {round(val_min*0.0725,2)}mm\n val_sqr : {val_sqr}px , {round(val_sqr*0.0725,2)}mm',
                            xy=(-5, 10), ha='center', va='baseline', fontsize=20)
        
        ## 그러주는건 여기까지.
        
        return res

 ###포인트 그래프 체크박스           
    def draw_round1(self,state):
        
        try:
            if state == qt.Checked:
                if 'csv' in self.Round[1].text():
                    self.df = []
                    self.df =self.df1
                    self.combo_fun()
                    res = self.maindraw(r_w=1)


                    res_aver = res/len(self.aver_logic) #평균좌표

                    # #로봇 움직임 표시
                    points = []
                    for i in range(0,46): ## 포인트 세팅. 45까지 있어야함
                        points.append(f"Point{i+1}")
                    for i in points:
                        if self.combo_list.currentText() == i:
                            number = points.index(i)+1
                            
                            #'왼쪽으로 이동'
                            if number < 7 or 13 < number < 19 or number==22 or 30<number<35 or 40<number<44:
                                points_move = -1 
                            
                            #'아래로 이동'
                            if number==7 or number==13 or number==19 or number==21 or number==23 or number==30: 
                                points_move = 0
                            
                            #'오른쪽으로 이동' 
                            if 7 < number < 13 or number==20 or 24<number<30 or 35<number<40 or number==44 or 43<number<47:
                                points_move = 1
                            
                            #대각 오른쪽 밑
                            if number==24 or number==35: #14~18포인트
                                points_move = 24
                            #대각 왼쪽 밑
                            if number==40: #14~18포인트
                                points_move = 40#'왼쪽으로 이동'


                    if res_aver>0:
                        if points_move ==-1:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n<----------------------------')
                        if points_move ==0:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n|\n|\n|\n|\n|\n|\n|\n|\n|\n|\n\\/')
                        if points_move ==1:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n---------------------------->')
                        if points_move == 24:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n\\\n \\\n  \\\n   \\\n    \\\n     \\\n      \\\n        _|')
                        if points_move == 40:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n          /\n         /\n        /\n       /\n      /\n     /\n    /\n   /\n\\  /\n')
                    if res_aver<0:
                        if points_move ==-1:
                            self.sc.axes.text((res_aver),1,'Slow Point\nRobot Movement\n<----------------------------')
                        if points_move ==0:
                            self.sc.axes.text((res_aver),1,'Slow Point\nRobot Movement\n|\n|\n|\n|\n|\n|\n|\n\\/')
                        if points_move ==1:
                            self.sc.axes.text((res_aver),1,'Slow Point\nRobot Movement\n---------------------------->')
                        if points_move == 24:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n\\\n \\\n  \\\n   \\\n    \\\n     \\\n      \\\n        _|')
                        if points_move == 40:
                            self.sc.axes.text((res_aver),1,'Fast Point\nRobot Movement\n          /\n         /\n        /\n       /\n      /\n     /\n    /\n   /\n\\  /\n')
                    
 
                    self.auto_basic_line()
                    self.sc.draw()
                else:
                    QMessageBox.warning(self, "Error code 01", "No Data")

            if state != qt.Checked:
                self.unchecked()
        except:pass
        
        
    def draw_round2(self,state):
        try:
            if state == qt.Checked:
                if 'csv' in self.Round[2].text():
                    self.df = []
                    self.df=self.df2
                    self.combo_fun()
                    self.maindraw(r_w=2)
                    self.auto_basic_line()
                    self.sc.draw()
                else:
                    QMessageBox.warning(self, "Error code 01", "No Data")

            if state != qt.Checked:
                self.unchecked()
        except:pass
    def draw_round3(self,state):
        try:
            if state == qt.Checked:
                if 'csv' in self.Round[3].text():
                    self.df = []
                    self.df =self.df3
                    self.combo_fun()
                    self.maindraw(r_w=3)
                    self.auto_basic_line()
                    self.sc.draw()
                else:
                    QMessageBox.warning(self, "Error code 01", "No Data")

            if state != qt.Checked:
                self.unchecked()
        except:pass
    def draw_round4(self,state):
        try:
            if state == qt.Checked:
                if 'csv' in self.Round[4].text():
                    self.df = self.df4
                    self.combo_fun()
                    self.maindraw(r_w=4)
                    self.auto_basic_line()
                    self.sc.draw()
                else:
                    QMessageBox.warning(self, "Error code 01", "No Data")

            if state != qt.Checked:
                self.unchecked()
        except:pass
    def draw_round5(self,state):
        try:
            if state == qt.Checked:
                if 'csv' in self.Round[5].text():
                    self.df =self.df5
                    self.combo_fun()
                    self.maindraw(r_w=5)
                    self.auto_basic_line()
                    self.sc.draw()
                else:
                    QMessageBox.warning(self, "Error code 01", "No Data")

            if state != qt.Checked:
                self.unchecked()
        except:pass
        
    def auto_basic_line(self):
            ####레퍼런스 포인트
        self.sc.axes.hlines(0, -30, 30,  linestyle='--', linewidth=1,color='red') #hlines 수평선
        self.sc.axes.vlines(self.set_reference_x-1500, -30, 30,  linestyle='--', linewidth=1,color='red') #수직 (x)
        draw_circle = plt.Circle((self.set_reference_x-1500, 0), 10, color='red',linewidth=1,fill=False)
        self.sc.axes.add_artist(draw_circle)
        self.sc.draw()
        
    def mains(self,df,r_how):
        # df = self.df1
        data = []
        dev_data = []
        for i in range(0,len(df)):
            now_point=df.loc[i] # 포인트
            re_x = now_point.iloc[0:1] #레퍼런스 포인트
            target_x = now_point[1:]
            #타겟 1부터 끝 타겟까지######################
            try: ## 평균값 옮기기
                for it in range(0,len(target_x)+1):
                    if target_x[it] != 0:
                        data.append(round(target_x[it]-1500,2))
                    if target_x[it] == 0:
                        data.append(target_x[it])
            except:continue
        

        if r_how == 2:
            self.R_Creat_[1].setText('Devitaion 1')
            self.data1 = abs(sum(data))/len(df) ###절대값
            self.sc.axes.plot([0, r_how], [0, (abs(sum(data))/len(df))])
            self.sc.axes.text(0,self.data1+0.3, f'Aver_Sub = {round(self.data1,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 4:
            self.R_Creat_[2].setText('Devitaion 2')
            self.data2 = abs(sum(data))/len(df)
            self.sc.axes.plot([2, r_how], [self.data1, (abs(sum(data))/46)])
            self.sc.axes.text(2,self.data2+0.3, f'Aver_Sub = {round(self.data2,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 6:
            self.R_Creat_[3].setText('Devitaion 3')
            self.data3 = abs(sum(data))/len(df)
            self.sc.axes.plot([4, r_how], [self.data2, (abs(sum(data))/46)])
            self.sc.axes.text(4,self.data3+0.3, f'Aver_Sub = {round(self.data3,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 8:
            self.R_Creat_[4].setText('Devitaion 4')
            self.data4= abs(sum(data))/len(df)
            self.sc.axes.plot([6, r_how], [self.data3, (abs(sum(data))/46)])
            self.sc.axes.text(6,self.data4+0.3, f'Aver_Sub = {round(self.data4,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 10:
            self.R_Creat_[5].setText('Devitaion 5')
            self.data5 = abs(sum(data))/len(df)
            self.sc.axes.plot([8, r_how], [self.data4, (abs(sum(data))/46)])
            self.sc.axes.text(8,self.data5+0.3, f'Aver_Sub = {round(self.data5,2)}',fontdict={'size': 8}) #산점도 레이블

            
        res = []
        far_pix = []
        res_num = 0
        for i in range(1,len(data)+1):
            try:
                if data[i-1] != -1500:
                    res.append(data[i-1])
                    res_num = res_num+1
                if i%5 == 0:
                    dev_data.append(round(np.std(res)*0.0725,2))
                    far_pix.append(round((sum(res)/res_num)*0.0725,2))
                    if r_how ==2:
                        if round(np.std(res)*0.0725,2) > 0.5:
                            self.R_Creat_[1].append(f'<p style="color: orange">Point{len(dev_data)} = {str(round(np.std(res)*0.0725,2))}</p>')
                        else:self.R_Creat_[1].append(f'Point{len(dev_data)} = '+str(round(np.std(res)*0.0725,2)))
                    if r_how ==4:
                        if round(np.std(res)*0.0725,2) > 0.5:
                            self.R_Creat_[2].append(f'<p style="color: orange">Point{len(dev_data)} = {str(round(np.std(res)*0.0725,2))}</p>')
                        else:self.R_Creat_[2].append(f'Point{len(dev_data)} = '+str(round(np.std(res)*0.0725,2)))
                    if r_how ==6:
                        if round(np.std(res)*0.0725,2) > 0.5:
                            self.R_Creat_[3].append(f'<p style="color: orange">Point{len(dev_data)} = {str(round(np.std(res)*0.0725,2))}</p>')
                        else:self.R_Creat_[3].append(f'Point{len(dev_data)} = '+str(round(np.std(res)*0.0725,2)))
                    if r_how ==8:
                        if round(np.std(res)*0.0725,2) > 0.5:
                            self.R_Creat_[4].append(f'<p style="color: orange">Point{len(dev_data)} = {str(round(np.std(res)*0.0725,2))}</p>')
                        else:self.R_Creat_[4].append(f'Point{len(dev_data)} = '+str(round(np.std(res)*0.0725,2)))
                    if r_how ==10:
                        if round(np.std(res)*0.0725,2) > 0.5:
                            self.R_Creat_[5].append(f'<p style="color: orange">Point{len(dev_data)} = {str(round(np.std(res)*0.0725,2))}</p>')
                        else:self.R_Creat_[5].append(f'Point{len(dev_data)} = '+str(round(np.std(res)*0.0725,2)))
                       
                    res = []
                    res_num = 0
            except:continue
        
        self.sc.axes.scatter(r_how,abs(sum(data))/len(df),c='black',s=100) #평균점

        if r_how == 2:
            self.dev_data1 = round(sum(dev_data)/len(df),2)
            self.R_Creat_[1].append(f'<p style="color: blue">Aver_dev = {str(round(sum(dev_data)/46,2))}</p>')
            self.sc.axes.plot([0, r_how], [0, (round(sum(dev_data)/46,2))])
            self.sc.axes.text(0,-0.4, f'Aver_Dev = {round(sum(dev_data)/46,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 4:
            self.dev_data2 = round(sum(dev_data)/len(df),2)
            self.R_Creat_[2].append(f'<p style="color: blue">Aver_dev = {str(round(sum(dev_data)/46,2))}</p>')
            self.sc.axes.plot([2, r_how], [self.dev_data1, (round(sum(dev_data)/46,2))])
            self.sc.axes.text(2,-0.8, f'Aver_Dev = {round(sum(dev_data)/46,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 6:
            self.dev_data3 = round(sum(dev_data)/len(df),2)
            self.R_Creat_[3].append(f'<p style="color: blue">Aver_dev = {str(round(sum(dev_data)/46,2))}</p>')
            self.sc.axes.plot([4, r_how], [self.dev_data2, (round(sum(dev_data)/46,2))])
            self.sc.axes.text(4,-1, f'Aver_Dev = {round(sum(dev_data)/46,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 8:
            self.dev_data4 = round(sum(dev_data)/len(df),2)
            self.R_Creat_[4].append(f'<p style="color: blue">Aver_dev = {str(round(sum(dev_data)/46,2))}</p>')
            self.sc.axes.plot([6, r_how], [self.dev_data3, (round(sum(dev_data)/46,2))])
            self.sc.axes.text(6,-1.2, f'Aver_Dev = {round(sum(dev_data)/46,2)}',fontdict={'size': 8}) #산점도 레이블
        if r_how == 10:
            self.dev_data5 = round(sum(dev_data)/len(df),2)
            self.R_Creat_[5].append(f'<p style="color: blue">Aver_dev = {str(round(sum(dev_data)/46,2))}</p>')
            self.sc.axes.plot([8, r_how], [self.dev_data4, (round(sum(dev_data)/46,2))])
            self.sc.axes.text(8,-1.4, f'Aver_Dev = {round(sum(dev_data)/46,2)}',fontdict={'size': 8}) #산점도 레이블
        
        self.sc.draw()
###################
        #print(far_pix)
        print(data)
        if r_how != 2:
            self.data_logic(df,r_how,data)
        elif r_how == 2:
            self.data_write(df,r_how)

        

    def data_logic(self,df,r_how,data):
        if r_how == 4:
            temp = int(self.log_read_line(19))
        if r_how == 6:
            temp = int(self.log_read_line(20))
        if r_how == 8:
            temp = int(self.log_read_line(21))
        if r_how == 10:
            temp = int(self.log_read_line(22))
        
        i2 = []
        res_aver = []
        for i in range(0,data):
            i2.append(i)
            if len(i2)==5:
                res_aver.append(sum(i2))
                i2 = []

        recom_val = []
        temp2 = 0
        far_pix = round(-float(self.set_reference_x-1500-res_aver),2)

        with open(f'{self.nowdir}\\data_log{temp-1}.txt','r') as read_txt:
                txtlist = read_txt.readlines()
                far_pix = round(far_pix*0.0725,2)
                for i in txtlist:
                    if txtlist.index(i) >= 15 and txtlist.index(i)%4==3:
                        txt = txtlist[i]
                        txt.replace('TriggIO vision_on,')
                        txt.replace('\\DOp:=do01_VisionTrigg,1;\\n')
                        recom_val.append(txt)

        print(recom_val)

                        recom = float(txt)+far_pix
                        recom은 data_write로 이동해야함.
        self.data_write(far_pix,df,r_how)
  


        

    def r1_create(self):
        self.reset_check() # 포인트 그려진거 해제
        df = self.df1
        self.mains(df,2)

    def r2_create(self):
        self.reset_check()
        df = self.df2
        self.mains(df,4)

    def r3_create(self):
        self.reset_check()
        df = self.df3
        self.mains(df,6)
    def r4_create(self):
        self.reset_check()
        df = self.df4
        self.mains(df,8)
    def r5_create(self):
        self.reset_check()
        df = self.df5
        self.mains(df,10)

    def open_r1(self):
        if self.create_number==0:
            QMessageBox.warning(self, "Error code 03", "No data")
        if self.create_number>0:
            # temp = self.create_number-1
            temp = int(self.log_read_line(19))
        subprocess.Popen(f'{self.nowdir}\\data_log{temp-1}.txt', shell=True)


    def open_r2(self):
        if self.create_number==0:
            QMessageBox.warning(self, "Error code 03", "No data")
        if self.create_number>0:
            # temp = self.create_number-1
            temp = int(self.log_read_line(20))
        subprocess.Popen(f'{self.nowdir}\\data_log{temp-1}.txt', shell=True)


    def open_r3(self):
        if self.create_number==0:
            QMessageBox.warning(self, "Error code 03", "No data")
        if self.create_number>0:
            # temp = self.create_number-1
            temp = int(self.log_read_line(21))
        subprocess.Popen(f'{self.nowdir}\\data_log{temp-1}.txt', shell=True)


    def open_r4(self):
        if self.create_number==0:
            QMessageBox.warning(self, "Error code 03", "No data")
        if self.create_number>0:
            # temp = self.create_number-1
            temp = int(self.log_read_line(22))
        subprocess.Popen(f'{self.nowdir}\\data_log{temp-1}.txt', shell=True)


    def open_r5(self):
        if self.create_number==0:
            QMessageBox.warning(self, "Error code 03", "No data")
        if self.create_number>0:
            # temp = self.create_number-1
            temp = int(self.log_read_line(23))
        subprocess.Popen(f'{self.nowdir}\\data_log{temp-1}.txt', shell=True)


    def def_back(self):
        pass
    
    def modechange(self):
        if self.modech == 0:
            self.sc.axes.axis([-2, 12, -2, 12])
            self.sc.draw()
            self.modech = 1
        elif self.modech == 1:
            self.sc.axes.axis([-1, 12, -2, 1.5])
            self.sc.draw()
            self.modech = 2
        elif self.modech == 2:
            self.sc.axes.axis([-15, 15, -15, 15])
            self.sc.draw()
            self.modech = 0

    
    def unchecked(self):
        try:
            if self.r_btn[1].isChecked() == True:
                redraw1 = 1
        except:pass
        try:
            if self.r_btn[2].isChecked() == True:
                redraw2 = 1
        except:pass
        try:
            if self.r_btn[3].isChecked() == True:
                redraw3 = 1
        except:pass
        try:
            if self.r_btn[4].isChecked() == True:
                redraw4 = 1
        except:pass
        try:
            if self.r_btn[5].isChecked() == True:
                redraw5 = 1
        except:pass
        
        self.reset_check()
        self.sc.axes.cla()
        self.sc.axes.axis([-15, 15, -15, 15])
        self.sc.axes.hlines(0, -30, 30,  linestyle='-', linewidth=2) #hlines 수평선 (y)
        self.sc.axes.vlines(0, -30, 30,  linestyle='-', linewidth=2)
        self.sc.axes.grid()
        self.sc.draw()
        
        try:
            if redraw1 == 1:
                self.r_btn[1].setChecked(True)
        except:pass
        try:
            if redraw2 == 1:
                self.r_btn[2].setChecked(True)
        except:pass
        try:
            if redraw3 == 1:
                self.r_btn[3].setChecked(True)
        except:pass
        try:
            if redraw4 == 1:
                self.r_btn[4].setChecked(True)
        except:pass
        try:
            if redraw5 == 1:
                self.r_btn[5].setChecked(True)
        except:pass
        redraw1 = 0
        redraw2 = 0
        redraw3 = 0
        redraw4 = 0
        redraw5 = 0

    def init_loop(self):
        pass



    



if __name__ == '__main__':

    app = QApplication(sys.argv)
    main = Main()
    sys.exit(app.exec_())
